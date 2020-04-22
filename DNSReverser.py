import ipwhois, sys, csv, os, re, logging
from datetime import datetime


import xlsxwriter
# Import the email modules we'll need

from pprint import pprint
from collections import OrderedDict
from openpyxl import load_workbook


logfilename = str(datetime.now()).replace(":","_")+".log"
logging.basicConfig(filename=logfilename,
                    filemode='a',
                    format='%(asctime)s,  %(name)s %(levelname)s %(message)s',
                    datefmt='%x %H:%M:%S',
                    level=logging.DEBUG)
logger = logging.getLogger(__name__)


START_FROM_COL = 1



def atoi(text):
    return int(text) if text.isdigit() else text

def natural_keys(text):
    '''
    alist.sort(key=natural_keys) sorts in human order
    http://nedbatchelder.com/blog/200712/human_sorting.html
    (See Toothy's implementation in the comments)
    '''
    return [ atoi(c) for c in re.split('(\d+)', text) ]





class LookupEntry:
	def __init__(self, ipaddress, out_file, dns_name=None):
		self.ipaddress = ipaddress
		self.email = ""
		self.address = ""
		self.company_name = ""
		self.contact_name = ""
		self.contact_fax = ""
		self.contact_phone = ""
		self.description = ""
		self.predefined_dns_name = dns_name
		self.out_file = out_file
			
	def resolveIP(self):
		try:
			# after 3 retries ask Japan and Korea
			obj = ipwhois.IPWhois(self.ipaddress).lookup_rdap(retry_count=3, 
													asn_methods=['dns', 'whois'], inc_nir=True)
									
			self.results_to_dict(obj)
			
		except ipwhois.exceptions.IPDefinedError: 
			self.address ="Private IP address"
			self.company_name ="Private IP address"
			self.contact_name = "Private IP address"
			self.contact_phone = "Private IP address"
			self.country = "Private IP address"
			self.description = "Private IP address"
		except ipwhois.exceptions.HTTPLookupError as e:
			self.address = "HTTP Lookup Error"
			self.company_name = "HTTP Lookup Error"
			self.contact_name =  "HTTP Lookup Error"
			self.contact_phone = "HTTP Lookup Error"
			self.country =  "HTTP Lookup Error"
			self.description = "HTTP Lookup Error"
			logger.warning("HTTP Lookup Error at IP {}".format(self.ipaddress))
			print(e)
			
		except Exception as e:
			self.address = "General Exception"
			self.company_name = "General Exception"
			self.contact_name = "General Exception"
			self.contact_phone = "General Exception"
			self.country = "General Exception"
			self.description = "General Exception"
			logger.warning("general error {}".format(e))

		cached_IPs[self.ipaddress] = self.country,\
							self.company_name, self.contact_name, self.contact_phone, self.contact_fax, \
							self.email, self.address, self.description

	def results_to_dict(self, results):
		try:
			self.description =  results['network']['remarks'][0]['description'].replace("\n", " ")
		except TypeError:
			self.description = "-"
		except IndexError:
			self.description = "-"
			
		self.country = results['network']['country']
		self.company_name = results['network']['name']

		for key, val in results['objects'].items():

			if isinstance(val, dict) and val['contact']:
				if val['contact']['address']:
					try:
						self.address = val['contact']['address'][0]['value'].\
							replace("\n", " ")
					except UnicodeEncodeError:
						self.address = "encoding error"
						handleUnicodeError()
						continue
						
				if val['contact']['email']:
					try:
						self.email = val['contact']['email'][0]['value']
					except UnicodeEncodeError:
						self.email = "encoding error"
						handleUnicodeError()
						continue
				if val['contact']['name']:
					try:
						self.contact_name = val['contact']['name']
					except UnicodeEncodeError:
						self.contact_name = "encoding error"
						handleUnicodeError()
						continue
				
				if val['contact']['phone']:
					for dict_ in val['contact']['phone']:
						tuple1, tuple2 = dict_.items()
						_, number = tuple1 
						_, number_type = tuple2
							
						if number_type == 'voice':
							try:
								self.contact_phone = number
							except UnicodeEncodeError:
								self.contact_phone = "encoding error"
								handleUnicodeError()
								continue
						else:
							try:
								self.contact_fax = number
							except UnicodeEncodeError:
								self.contact_fax = "encoding error"
								handleUnicodeError()
								continue
	
		
def appendToCSVfile(ipaddress, out_file, timestamp=None):
	with open(out_file, "a") as csvwriter:
		writer = csv.writer(csvwriter, delimiter="|")
		logging.info("Writing data {}".format(ipaddress))
		writer.writerow([timestamp, ipaddress, *cached_IPs[ipaddress]])
	

		
def writeToXLSXfile(xlsxfile, ip_resolved_data):
		
	#logger.info("writing to xlsx file{}".format(xlsxfile)
	workbook = xlsxwriter.Workbook(xlsxfile)
	worksheet = workbook.add_worksheet("extracted data")
	worksheet.write_row(0, 0, ("time, provider", 
								"country", "provider email",
								 "provider address", "description", 
								 "predefined dns name"))
								
	for row_idx, (timestamp, name, country, email, 
		address, description, predefined_dns_name) in enumerate(ip_resolved_data):
		
		worksheet.write_row(row_idx+1, 0, (name, country, email, 
			address, description, predefined_dns_name))
	workbook.close()
      
                
        
class ReadIPs:
	
	def __init__(self, file_name, mode=None):
		self.file_name = file_name
	
	def __iter__(self):
		if self.file_name.endswith("xlsx"):
		
			wb = load_workbook(filename=self.file_name, read_only=True)
			ws = wb['Sheet1']

			for row_num, row in enumerate(ws.rows):
				if row_num >= START_FROM_COL:
					yield row_num, row[1].value, str(row[2].value).strip()
		elif self.file_name.endswith("txt"):
			with open(self.file_name, newline='') as txtfile:
				for row_num, row in enumerate(txtfile.readlines()):
					yield row_num, "", row.strip().replace("\n", "")
		
		elif mode == "Repass":
			with open(self.file_name, 'r', newline='') as csvfile:
				ipreader  = csv.reader(csvfile, delimiter='|')
				for row_num, row in enumerate(ipreader):
					if row_num >= START_FROM_COL:		
							yield row_num, row

		else:
			with open(self.file_name, newline='\r\n') as csvfile:
				ipreader  = csv.reader(csvfile, delimiter=',')
				for row_num, row in enumerate(ipreader):
					if row_num >= START_FROM_COL:
						yield row_num, row
      




def handleUnicodeError():
	logger.warning("encoding error with address")




def process_ip(provider_ip):

	if provider_ip not in cached_IPs.keys():
		lookupentry = LookupEntry(provider_ip, out_file)
		lookupentry.resolveIP()
		
		logger.info("resolved IP {}".format(provider_ip))
		
	else: 
		
		logger.info("using cached entry for IP {}".format(provider_ip))
	
	
	
	




if __name__ == "__main__":
	cached_IPs = {}
	ip_resolved_data = {}
	
	mode = None
	
	out_file = sys.argv[2]
	if os.path.exists(out_file):
		os.remove(out_file)
	
	
	with open(out_file, "a") as csvwriter:
		writer = csv.writer(csvwriter, delimiter="|")
		writer.writerow(["Date & time", "IP",  "Country", "Provider Name", "Contact Name",
				"phone", "fax", "e-mail", "address", "description"])
	
	print ("READING", sys.argv[1])
	
	if len(sys.argv) >2:
		mode = sys.argv[3]
	
	for row_num, vals in ReadIPs(sys.argv[1], mode):
		ip = vals[2]
		timestamp = vals[1]

		if not cached_IPs.get(ip):
			print ("Resolving IP {} from row {}".format(ip, row_num))
			
			process_ip(ip)

		appendToCSVfile(ip, out_file, timestamp)
		
		
		
		
	
	
	
		


